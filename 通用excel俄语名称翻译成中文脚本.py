import re
import openpyxl
import logging
import requests
import time
import os
import json
from getpass import getpass

# 配置日志
logging.basicConfig(level=logging.INFO,
                    format='%(asctime)s - %(levelname)s - %(message)s',
                    handlers=[
                        logging.FileHandler('translation.log', encoding='utf-8'),
                        logging.StreamHandler()
                    ])
logger = logging.getLogger()

def get_column_index(input_str):
    """将列字母（如A/B）或数字转换为数字索引"""
    try:
        # 尝试转换为数字
        col = int(input_str)
        if col < 1:
            raise ValueError("列号不能小于1")
        return col
    except ValueError:
        # 处理字母转数字
        col_letter = input_str.upper().strip()
        if not col_letter.isalpha():
            raise ValueError("无效的列标识")
        return openpyxl.utils.column_index_from_string(col_letter)

def get_user_config():
    """获取用户配置"""
    print("\n" + "="*40)
    print("【配置向导】")
    print("="*40)
    
    # 修改为可见的API密钥输入
    api_key = input("请输入DeepSeek API密钥（输入可见）：").strip()
    file_path = input("请输入Excel文件完整路径：").strip(' "\'')
    
    # 列配置
    while True:
        try:
            src_col = get_column_index(input("俄语内容所在列（例如A或1）："))
            break
        except Exception as e:
            print(f"错误：{str(e)}，请重新输入")

    while True:
        try:
            tgt_col = get_column_index(input("中文翻译存放列（例如B或2）："))
            break
        except Exception as e:
            print(f"错误：{str(e)}，请重新输入")

    return {
        'api_key': api_key,
        'file_path': file_path,
        'src_col': src_col,
        'tgt_col': tgt_col
    }

def clean_translation(text):
    """清洗翻译结果"""
    return re.sub(r'^\d+\.+\s*', '', text.strip())

def main():
    # 获取配置
    try:
        config = get_user_config()
    except Exception as e:
        logger.error(f"配置错误: {str(e)}")
        return

    # 初始化API参数
    DEEPSEEK_API_URL = 'https://api.deepseek.com/v1/chat/completions'
    headers = {
        'Content-Type': 'application/json',
        'Authorization': f'Bearer {config["api_key"]}'
    }

    workbook = None
    try:
        # 加载工作簿
        workbook = openpyxl.load_workbook(config['file_path'])
        sheet = workbook.active  # 使用当前活动工作表
        max_row = sheet.max_row
        logger.info(f"文件加载成功，总行数：{max_row}")
        logger.info(f"俄语列：{config['src_col']} → 中文列：{config['tgt_col']}")

        # 批量处理
        for batch_num, start_row in enumerate(range(1, max_row + 1, 100), 1):
            end_row = min(start_row + 99, max_row)
            batch_size = end_row - start_row + 1

            # 读取源文本
            sources = []
            for row in range(start_row, end_row + 1):
                cell = sheet.cell(row=row, column=config['src_col'])
                sources.append(str(cell.value) if cell.value else "")

            # 构建prompt
            prompt = (
                "请严格遵循以下要求：\n"
                "1. 逐行翻译下方俄语到简体中文\n"
                "2. 保持完全相同的行数\n"
                "3. 不要添加任何序号或标记\n"
                "4. 空行保持为空\n\n"
                "待翻译内容：\n" + '\n'.join(sources)
            )

            # API请求
            try:
                response = requests.post(
                    DEEPSEEK_API_URL,
                    json={
                        "model": "deepseek-chat",
                        "messages": [{"role": "user", "content": prompt}],
                        "temperature": 0.1
                    },
                    headers=headers,
                    timeout=30
                )
                response.raise_for_status()
                translated = response.json()['choices'][0]['message']['content'].split('\n')
            except Exception as e:
                logger.error(f"第 {batch_num} 批翻译失败: {str(e)}")
                translated = []

            # 处理结果
            cleaned = [clean_translation(t) for t in translated]
            if len(cleaned) < batch_size:
                cleaned += [''] * (batch_size - len(cleaned))
            elif len(cleaned) > batch_size:
                cleaned = cleaned[:batch_size]

            # 写入结果
            for idx, row in enumerate(range(start_row, end_row + 1)):
                sheet.cell(row=row, column=config['tgt_col'], value=cleaned[idx])

            # 实时保存
            workbook.save(config['file_path'])
            logger.info(f"批次 {batch_num}（行 {start_row}-{end_row}）完成，保存成功")

    except Exception as e:
        logger.error(f"程序异常: {str(e)}", exc_info=True)
    finally:
        if workbook:
            workbook.close()

if __name__ == "__main__":
    main()
