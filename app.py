import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import json
import os
import threading
import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
import translator
import logging
import time
from logging.handlers import RotatingFileHandler
import io
import sys
import requests

# --- 日志和字体配置 ---
class TextWidgetHandler(logging.Handler):
    def __init__(self, text_widget):
        super().__init__()
        self.text_widget = text_widget
        self.text_widget.tag_configure("INFO", foreground="black")
        self.text_widget.tag_configure("DEBUG", foreground="gray")
        self.text_widget.tag_configure("WARNING", foreground="orange")
        self.text_widget.tag_configure("ERROR", foreground="red")
        self.text_widget.tag_configure("CRITICAL", foreground="red", font=("Microsoft YaHei UI", 10, "bold"))

    def emit(self, record):
        msg = self.format(record)
        def append_log():
            try:
                self.text_widget.config(state=tk.NORMAL)
                self.text_widget.insert(tk.END, msg + "\n", record.levelname)
                self.text_widget.see(tk.END)
                self.text_widget.config(state=tk.DISABLED)
            except tk.TclError:
                pass
        if self.text_widget.winfo_exists():
            self.text_widget.after(0, append_log)

logger = logging.getLogger() 
logger.setLevel(logging.DEBUG)
if logger.hasHandlers(): logger.handlers.clear()
file_handler = RotatingFileHandler('translation.log', maxBytes=1024*1024*5, backupCount=5, encoding='utf-8')
file_formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
file_handler.setFormatter(file_formatter)
logger.addHandler(file_handler)
try:
    console_handler = logging.StreamHandler(io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8'))
    console_handler.setLevel(logging.INFO)
    console_handler.setFormatter(file_formatter)
    logger.addHandler(console_handler)
except Exception: pass
DEFAULT_FONT = ("Microsoft YaHei UI", 10)
HEADER_FONT = ("Microsoft YaHei UI", 10, "bold")


# --- ExcelPreview 类 (无变化) ---
class ExcelPreview(ttk.Frame):
    def __init__(self, parent, app_instance, *args, **kwargs):
        super().__init__(parent, *args, **kwargs)
        self.app_instance = app_instance
        self.canvas = tk.Canvas(self, bg="white", bd=2, relief="sunken")
        self.v_scroll = ttk.Scrollbar(self, orient="vertical", command=self.canvas.yview)
        self.h_scroll = ttk.Scrollbar(self, orient="horizontal", command=self.canvas.xview)
        self.canvas.configure(yscrollcommand=self.v_scroll.set, xscrollcommand=self.h_scroll.set)
        self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.v_scroll.pack(side=tk.RIGHT, fill="y")
        self.h_scroll.pack(side=tk.BOTTOM, fill="x")
        self.canvas.bind("<Configure>", self._on_canvas_configure)
        self.canvas.bind("<Button-1>", self._on_left_click)
        self.canvas.bind("<Button-3>", self._on_right_click)
        self.sheet = None
        self.cell_width = 100
        self.cell_height = 25
        self.header_height = 25
        self.row_header_width = 50
        self.selected_src_col = None
        self.selected_src_row = None
        self.selected_tgt_col = None
        self.drawn_rects = {}
    def _on_canvas_configure(self, event):
        self.canvas.config(scrollregion=self.canvas.bbox("all"))
        self.draw_sheet()
    def load_sheet(self, sheet):
        self.sheet = sheet
        self.selected_src_col = None
        self.selected_src_row = None
        self.selected_tgt_col = None
        self.draw_sheet()
        self.app_instance.update_selection_display()
    def draw_sheet(self):
        if not self.sheet: return
        self.canvas.delete("all")
        self.drawn_rects = {}
        max_row = self.sheet.max_row
        max_col = self.sheet.max_column
        for c_idx in range(1, max_col + 1):
            col_letter = openpyxl.utils.get_column_letter(c_idx)
            x1 = self.row_header_width + (c_idx - 1) * self.cell_width
            y1 = 0
            x2 = x1 + self.cell_width
            y2 = self.header_height
            self.canvas.create_rectangle(x1, y1, x2, y2, outline="gray", fill="#f0f0f0", tags=f"col_header_bg_{c_idx}")
            self.canvas.create_text(x1 + self.cell_width / 2, y1 + self.header_height / 2, text=col_letter, font=HEADER_FONT)
        for r_idx in range(1, max_row + 1):
            x1 = 0
            y1 = self.header_height + (r_idx - 1) * self.cell_height
            x2 = self.row_header_width
            y2 = y1 + self.cell_height
            self.canvas.create_rectangle(x1, y1, x2, y2, outline="gray", fill="#f0f0f0")
            self.canvas.create_text(x1 + self.row_header_width / 2, y1 + self.cell_height / 2, text=str(r_idx), font=HEADER_FONT)
        for r_idx in range(1, max_row + 1):
            for c_idx in range(1, max_col + 1):
                cell = self.sheet.cell(row=r_idx, column=c_idx)
                x1 = self.row_header_width + (c_idx - 1) * self.cell_width
                y1 = self.header_height + (r_idx - 1) * self.cell_height
                x2 = x1 + self.cell_width
                y2 = y1 + self.cell_height
                rect_id = self.canvas.create_rectangle(x1, y1, x2, y2, outline="gray", fill="white")
                self.drawn_rects[(r_idx, c_idx)] = rect_id
                display_value = cell.value
                for merged_range in self.sheet.merged_cells.ranges:
                    if cell.coordinate in merged_range and cell.coordinate != merged_range.coord.split(':')[0]:
                        display_value = ""
                        break
                self.canvas.create_text(x1 + 5, y1 + self.cell_height / 2, text=str(display_value) if display_value is not None else "", anchor="w", font=DEFAULT_FONT)
        self.canvas.config(scrollregion=(0, 0, self.row_header_width + max_col * self.cell_width, self.header_height + max_row * self.cell_height))
        self._highlight_selections()
    def _get_cell_from_coords(self, x, y):
        canvas_x = self.canvas.canvasx(x)
        canvas_y = self.canvas.canvasy(y)
        if canvas_x < self.row_header_width or canvas_y < self.header_height: return None, None
        col = int((canvas_x - self.row_header_width) / self.cell_width) + 1
        row = int((canvas_y - self.header_height) / self.cell_height) + 1
        if not self.sheet or col > self.sheet.max_column or row > self.sheet.max_row: return None, None
        return row, col
    def _on_left_click(self, event):
        row, col = self._get_cell_from_coords(event.x, event.y)
        if row is not None and col is not None:
            self.selected_src_col = col
            self.selected_src_row = row
            self._highlight_selections()
            self.app_instance.update_selection_display()
    def _on_right_click(self, event):
        row, col = self._get_cell_from_coords(event.x, event.y)
        if col is not None:
            self.selected_tgt_col = col
            self._highlight_selections()
            self.app_instance.update_selection_display()
        return "break"
    def _highlight_selections(self):
        for (r, c), rect_id in self.drawn_rects.items(): self.canvas.itemconfig(rect_id, fill="white")
        if self.sheet:
            for c_idx in range(1, self.sheet.max_column + 1):
                col_header_bg_id = self.canvas.find_withtag(f"col_header_bg_{c_idx}")
                if col_header_bg_id: self.canvas.itemconfig(col_header_bg_id, fill="#f0f0f0")
        if self.selected_tgt_col and self.sheet:
            for r_idx in range(1, self.sheet.max_row + 1):
                rect_id = self.drawn_rects.get((r_idx, self.selected_tgt_col))
                if rect_id: self.canvas.itemconfig(rect_id, fill="#e0e0ff")
            col_header_bg_id = self.canvas.find_withtag(f"col_header_bg_{self.selected_tgt_col}")
            if col_header_bg_id: self.canvas.itemconfig(col_header_bg_id, fill="#c0c0ff")
        if self.selected_src_col and self.sheet:
            for r_idx in range(1, self.sheet.max_row + 1):
                rect_id = self.drawn_rects.get((r_idx, self.selected_src_col))
                if rect_id: self.canvas.itemconfig(rect_id, fill="#e0ffe0")
            col_header_bg_id = self.canvas.find_withtag(f"col_header_bg_{self.selected_src_col}")
            if col_header_bg_id: self.canvas.itemconfig(col_header_bg_id, fill="#c0ffc0")
        if self.selected_src_col and self.selected_src_row:
            rect_id = self.drawn_rects.get((self.selected_src_row, self.selected_src_col))
            if rect_id: self.canvas.itemconfig(rect_id, fill="#66ff66", outline='red', width=2)
    def get_selected_source_coords(self):
        if self.selected_src_col and self.selected_src_row: return openpyxl.utils.get_column_letter(self.selected_src_col), self.selected_src_row
        return None, None
    def get_selected_target_col(self):
        if self.selected_tgt_col: return openpyxl.utils.get_column_letter(self.selected_tgt_col)
        return None
    def set_selected_source_coords(self, col_letter, row):
        if col_letter and row:
            try:
                self.selected_src_col = openpyxl.utils.column_index_from_string(col_letter)
                self.selected_src_row = int(row)
                self._highlight_selections()
            except (ValueError, TypeError): pass
    def set_selected_target_col(self, col_letter):
        if col_letter:
            try:
                self.selected_tgt_col = openpyxl.utils.column_index_from_string(col_letter)
                self._highlight_selections()
            except ValueError: pass

# --- ProxyManagerWindow 类 (无变化) ---
class ProxyManagerWindow(tk.Toplevel):
    def __init__(self, parent, app_instance):
        super().__init__(parent)
        self.title("代理管理")
        self.geometry("800x600")
        self.app_instance = app_instance
        self.transient(parent)
        self.grab_set()
        self.proxies = self.app_instance.proxies
        self.proxy_name_var = tk.StringVar()
        self.proxy_type_var = tk.StringVar(value="HTTP")
        self.proxy_address_var = tk.StringVar()
        self.proxy_port_var = tk.StringVar()
        self.proxy_user_var = tk.StringVar()
        self.proxy_pass_var = tk.StringVar()
        self._create_widgets()
        self._load_proxies_to_treeview()
        self.protocol("WM_DELETE_WINDOW", self._on_closing)

    def _create_widgets(self):
        input_frame = ttk.LabelFrame(self, text="代理详情", padding="10")
        input_frame.pack(padx=10, pady=10, fill=tk.X)
        ttk.Label(input_frame, text="代理名称:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=2)
        ttk.Entry(input_frame, textvariable=self.proxy_name_var).grid(row=0, column=1, sticky=tk.EW, padx=5, pady=2)
        ttk.Label(input_frame, text="代理类型:").grid(row=0, column=2, sticky=tk.W, padx=5, pady=2)
        ttk.Combobox(input_frame, textvariable=self.proxy_type_var, values=["HTTP", "SOCKS5"], state="readonly").grid(row=0, column=3, sticky=tk.EW, padx=5, pady=2)
        ttk.Label(input_frame, text="地址:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=2)
        ttk.Entry(input_frame, textvariable=self.proxy_address_var).grid(row=1, column=1, sticky=tk.EW, padx=5, pady=2)
        ttk.Label(input_frame, text="端口:").grid(row=1, column=2, sticky=tk.W, padx=5, pady=2)
        ttk.Entry(input_frame, textvariable=self.proxy_port_var).grid(row=1, column=3, sticky=tk.EW, padx=5, pady=2)
        ttk.Label(input_frame, text="用户名 (可选):").grid(row=2, column=0, sticky=tk.W, padx=5, pady=2)
        ttk.Entry(input_frame, textvariable=self.proxy_user_var).grid(row=2, column=1, sticky=tk.EW, padx=5, pady=2)
        ttk.Label(input_frame, text="密码 (可选):").grid(row=2, column=2, sticky=tk.W, padx=5, pady=2)
        ttk.Entry(input_frame, textvariable=self.proxy_pass_var, show="*").grid(row=2, column=3, sticky=tk.EW, padx=5, pady=2)
        input_frame.columnconfigure(1, weight=1)
        input_frame.columnconfigure(3, weight=1)
        button_frame = ttk.Frame(self)
        button_frame.pack(pady=10)
        self.test_button = ttk.Button(button_frame, text="测试连接", command=self._test_proxy)
        self.test_button.pack(side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="添加", command=self._add_proxy).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="更新选中", command=self._update_proxy).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="删除选中", command=self._delete_proxy).pack(side=tk.LEFT, padx=5)
        list_frame = ttk.LabelFrame(self, text="已配置代理", padding="10")
        list_frame.pack(padx=10, pady=5, fill=tk.BOTH, expand=True)
        self.tree = ttk.Treeview(list_frame, columns=("Type", "Address"), show="headings")
        self.tree.heading("#0", text="代理名称")
        self.tree.column("#0", width=150, stretch=tk.NO)
        self.tree.heading("Type", text="类型")
        self.tree.column("Type", width=80, anchor='center')
        self.tree.heading("Address", text="地址:端口")
        self.tree.column("Address", width=250)
        self.tree.pack(fill=tk.BOTH, expand=True, side=tk.LEFT)
        self.tree.bind("<<TreeviewSelect>>", self._on_tree_select)
        tree_scroll = ttk.Scrollbar(list_frame, orient="vertical", command=self.tree.yview)
        tree_scroll.pack(side=tk.RIGHT, fill="y")
        self.tree.configure(yscrollcommand=tree_scroll.set)
        ttk.Button(self, text="关闭窗口", command=self._on_closing).pack(pady=10)
        
    def _test_proxy(self):
        name, details = self._get_details_from_form()
        if not name:
            return
        self.test_button.config(state=tk.DISABLED)
        threading.Thread(target=self._proxy_test_worker, args=(details,), daemon=True).start()

    def _proxy_test_worker(self, details):
        proxy_type = details.get("type", "http").lower()
        address = details.get("address")
        port = details.get("port")
        username = details.get("username")
        password = details.get("password")
        auth = f"{username}:{password}@" if username and password else ""
        
        if proxy_type.startswith('socks'):
            try:
                import socks
            except ImportError:
                messagebox.showerror("依赖缺失", "检测到SOCKS代理，但缺少'PySocks'库.\n请在终端运行 'pip install PySocks' 来安装它.", parent=self)
                self.after(0, lambda: self.test_button.config(state=tk.NORMAL))
                return

        proxy_url = f"{proxy_type}://{auth}{address}:{port}"
        proxies = {"http": proxy_url, "https": proxy_url}
        test_url = "http://www.google.com/generate_204"
        
        try:
            response = requests.get(test_url, proxies=proxies, timeout=10)
            if response.status_code == 204:
                messagebox.showinfo("成功", "代理连接成功！", parent=self)
            else:
                messagebox.showwarning("警告", f"代理似乎工作正常，但测试网址返回状态码 {response.status_code}。", parent=self)
        except requests.exceptions.ProxyError as e:
            messagebox.showerror("错误", f"代理连接失败: \n{e}", parent=self)
        except requests.exceptions.Timeout:
            messagebox.showerror("错误", "连接超时。请检查代理地址、端口或您的网络。", parent=self)
        except Exception as e:
            messagebox.showerror("未知错误", f"测试代理时发生错误: \n{e}", parent=self)
        finally:
            self.after(0, lambda: self.test_button.config(state=tk.NORMAL))


    def _load_proxies_to_treeview(self):
        for i in self.tree.get_children(): self.tree.delete(i)
        for name, details in self.proxies.items():
            addr_port = f"{details.get('address', '')}:{details.get('port', '')}"
            self.tree.insert("", tk.END, iid=name, text=name, values=(details.get("type", "HTTP"), addr_port))
    def _on_tree_select(self, event=None):
        selected = self.tree.selection()
        if not selected: return
        name = selected[0]
        details = self.proxies.get(name)
        if details:
            self.proxy_name_var.set(name)
            self.proxy_type_var.set(details.get("type", "HTTP"))
            self.proxy_address_var.set(details.get("address", ""))
            self.proxy_port_var.set(details.get("port", ""))
            self.proxy_user_var.set(details.get("username", ""))
            self.proxy_pass_var.set(details.get("password", ""))
    def _get_details_from_form(self):
        name = self.proxy_name_var.get().strip()
        addr = self.proxy_address_var.get().strip()
        port = self.proxy_port_var.get().strip()
        if not name or not addr or not port:
            messagebox.showerror("错误", "代理名称、地址和端口不能为空！", parent=self)
            return None, None
        details = {"type": self.proxy_type_var.get(), "address": addr, "port": port, "username": self.proxy_user_var.get().strip(), "password": self.proxy_pass_var.get()}
        return name, details
    def _add_proxy(self):
        name, details = self._get_details_from_form()
        if not name: return
        if name in self.proxies:
            messagebox.showerror("错误", f"名为 '{name}' 的代理已存在！", parent=self)
            return
        self.proxies[name] = details
        self._save_and_update()
    def _update_proxy(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showerror("错误", "请选择要更新的代理！", parent=self)
            return
        original_name = selected[0]
        name, details = self._get_details_from_form()
        if not name: return
        if original_name != name and name in self.proxies:
            messagebox.showerror("错误", f"新的代理名称 '{name}' 已存在！", parent=self)
            return
        if original_name != name:
            del self.proxies[original_name]
        self.proxies[name] = details
        self._save_and_update()
        if self.app_instance.current_proxy_name_var.get() == original_name:
            self.app_instance.current_proxy_name_var.set(name)
    def _delete_proxy(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showerror("错误", "请选择要删除的代理！", parent=self)
            return
        name = selected[0]
        if messagebox.askyesno("确认删除", f"确定要删除代理 '{name}' 吗？", parent=self):
            del self.proxies[name]
            self._save_and_update()
            if self.app_instance.current_proxy_name_var.get() == name:
                self.app_instance.current_proxy_name_var.set("无代理")
    def _save_and_update(self):
        self._load_proxies_to_treeview()
        self.app_instance.save_config(self.app_instance.config_file)
        self.app_instance.update_proxy_combobox()
    def _on_closing(self):
        self.grab_release()
        self.destroy()

# --- ModelManagerWindow 类 (已重构) ---
class ModelManagerWindow(tk.Toplevel):
    def __init__(self, parent, app_instance):
        super().__init__(parent)
        self.title("AI模型管理")
        self.geometry("800x600")
        self.app_instance = app_instance
        self.transient(parent)
        self.grab_set()

        self.models = self.app_instance.models

        self.model_name_var = tk.StringVar()
        self.api_key_var = tk.StringVar()
        self.api_provider_var = tk.StringVar(value="Gemini")
        
        self.model_id_var = tk.StringVar()
        self.custom_api_url_var = tk.StringVar()

        self._create_widgets()
        self._load_models_to_treeview()
        self.protocol("WM_DELETE_WINDOW", self._on_closing)

    def _create_widgets(self):
        common_frame = ttk.LabelFrame(self, text="通用配置", padding="10")
        common_frame.pack(padx=10, pady=10, fill=tk.X)
        
        ttk.Label(common_frame, text="配置名称:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        ttk.Entry(common_frame, textvariable=self.model_name_var).grid(row=0, column=1, sticky=tk.EW, padx=5)

        ttk.Label(common_frame, text="API密钥:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        ttk.Entry(common_frame, textvariable=self.api_key_var, show="*").grid(row=1, column=1, sticky=tk.EW, padx=5)
        
        ttk.Label(common_frame, text="API提供商:").grid(row=2, column=0, sticky=tk.W, padx=5, pady=5)
        provider_combo = ttk.Combobox(common_frame, textvariable=self.api_provider_var, values=["Gemini", "DeepSeek", "Custom"], state="readonly")
        provider_combo.grid(row=2, column=1, sticky=tk.EW, padx=5)
        provider_combo.bind("<<ComboboxSelected>>", self._toggle_provider_fields)
        
        common_frame.columnconfigure(1, weight=1)

        self.provider_frame = ttk.Frame(self)
        self.provider_frame.pack(padx=10, pady=5, fill=tk.X, expand=True)

        self.model_config_frame = ttk.LabelFrame(self.provider_frame, text="模型配置", padding="10")
        self.model_config_frame.pack(fill=tk.X, expand=True, padx=10, pady=5)
        self.model_config_frame.columnconfigure(1, weight=1)

        ttk.Label(self.model_config_frame, text="选择模型:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        self.model_combo = ttk.Combobox(self.model_config_frame, textvariable=self.model_id_var, state="readonly")
        self.model_combo.grid(row=0, column=1, sticky=tk.EW, padx=5)
        self.fetch_models_button = ttk.Button(self.model_config_frame, text="获取模型列表", command=self._fetch_models_thread)
        self.fetch_models_button.grid(row=0, column=2, padx=5)

        self.custom_url_frame = ttk.LabelFrame(self.provider_frame, text="自定义API地址", padding="10")
        self.custom_url_frame.columnconfigure(1, weight=1)
        ttk.Label(self.custom_url_frame, text="API地址:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        ttk.Entry(self.custom_url_frame, textvariable=self.custom_api_url_var).grid(row=0, column=1, sticky=tk.EW, padx=5)
        
        button_frame = ttk.Frame(self)
        button_frame.pack(pady=10)
        ttk.Button(button_frame, text="添加", command=self._add_model).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="更新选中", command=self._update_model).pack(side=tk.LEFT, padx=5)
        ttk.Button(button_frame, text="删除选中", command=self._delete_model).pack(side=tk.LEFT, padx=5)

        list_frame = ttk.LabelFrame(self, text="已配置模型", padding="10")
        list_frame.pack(padx=10, pady=5, fill=tk.BOTH, expand=True)
        self.tree = ttk.Treeview(list_frame, columns=("Provider", "Details"), show="headings")
        self.tree.heading("#0", text="配置名称")
        self.tree.column("#0", width=150, stretch=tk.NO)
        self.tree.heading("Provider", text="提供商")
        self.tree.column("Provider", width=80, anchor='center')
        self.tree.heading("Details", text="模型ID / 地址")
        self.tree.column("Details", width=350)
        self.tree.pack(fill=tk.BOTH, expand=True, side=tk.LEFT)
        self.tree.bind("<<TreeviewSelect>>", self._on_tree_select)
        tree_scroll = ttk.Scrollbar(list_frame, orient="vertical", command=self.tree.yview)
        tree_scroll.pack(side=tk.RIGHT, fill="y")
        self.tree.configure(yscrollcommand=tree_scroll.set)
        
        ttk.Button(self, text="关闭窗口", command=self._on_closing).pack(pady=10)
        
        self._toggle_provider_fields()

    def _fetch_models_thread(self):
        api_key = self.api_key_var.get().strip()
        provider = self.api_provider_var.get()
        if not api_key:
            messagebox.showerror("错误", f"请先输入您的 {provider} API密钥。", parent=self)
            return
        self.fetch_models_button.config(state=tk.DISABLED, text="获取中...")
        threading.Thread(target=self._fetch_models_worker, args=(api_key, provider), daemon=True).start()

    def _fetch_models_worker(self, api_key, provider):
        try:
            proxy_name = self.app_instance.current_proxy_name_var.get()
            proxy_config = self.app_instance.proxies.get(proxy_name) if proxy_name != "无代理" else None
            
            model_list = []
            logger.debug(f"_fetch_models_worker: 传递的代理配置为: {proxy_config}")
            if provider == "Gemini":
                model_list = translator.fetch_gemini_models(api_key, proxy_config)
            elif provider == "DeepSeek":
                model_list = translator.fetch_deepseek_models(api_key, proxy_config)

            def update_ui_success():
                if model_list:
                    self.model_combo['values'] = model_list
                    self.model_id_var.set(model_list[0])
                    messagebox.showinfo("成功", f"成功获取到 {len(model_list)} 个模型。", parent=self)
                else:
                    self.model_combo['values'] = []
                    self.model_id_var.set("")
                    messagebox.showwarning("警告", f"未能获取到任何 {provider} 模型列表。", parent=self)
            self.after(0, update_ui_success)

        except Exception as e:
            logger.error(f"获取 {provider} 模型列表失败: {e}")
            def update_ui_error():
                 messagebox.showerror("错误", f"获取模型列表失败:\n{e}", parent=self)
            self.after(0, update_ui_error)
        finally:
            def reenable_button():
                self.fetch_models_button.config(state=tk.NORMAL, text="获取模型列表")
            self.after(0, reenable_button)

    def _toggle_provider_fields(self, event=None):
        provider = self.api_provider_var.get()
        if provider == "Custom":
            self.model_config_frame.pack(fill=tk.X, expand=True, padx=10, pady=5)
            self.custom_url_frame.pack(fill=tk.X, expand=True, padx=10, pady=5)
            self.fetch_models_button.grid_remove()
            self.model_combo.config(state="normal")
            self.model_combo.config(textvariable=self.model_id_var)
        else:
            self.model_config_frame.pack(fill=tk.X, expand=True, padx=10, pady=5)
            self.custom_url_frame.pack_forget()
            self.fetch_models_button.grid()
            self.model_combo.config(state="readonly")

    def _load_models_to_treeview(self):
        for i in self.tree.get_children(): self.tree.delete(i)
        for name, details in self.models.items():
            provider = details.get("provider", "Custom")
            detail_text = details.get("model_id")
            if provider == "Custom":
                detail_text = f"{details.get('api_url')} ({details.get('model_id')})"
            self.tree.insert("", tk.END, iid=name, text=name, values=(provider, detail_text))
            
    def _on_tree_select(self, event=None):
        selected = self.tree.selection()
        if not selected: return
        name = selected[0]
        details = self.models.get(name)
        if details:
            self.model_name_var.set(name)
            self.api_key_var.set(details.get("api_key", ""))
            provider = details.get("provider", "Custom")
            self.api_provider_var.set(provider)
            
            saved_model_id = details.get("model_id", "")
            self.model_combo['values'] = [saved_model_id] if saved_model_id else []
            self.model_id_var.set(saved_model_id)
            
            self.custom_api_url_var.set(details.get("api_url", ""))
        self._toggle_provider_fields()

    def _collect_and_validate(self):
        name = self.model_name_var.get().strip()
        api_key = self.api_key_var.get().strip()
        provider = self.api_provider_var.get()
        if not name or not api_key:
            messagebox.showerror("错误", "配置名称和API密钥不能为空！", parent=self)
            return None
        
        details = {"provider": provider, "api_key": api_key}
        model_id = self.model_id_var.get().strip()
        if not model_id:
            messagebox.showerror("错误", "模型ID不能为空！", parent=self)
            return None
        details["model_id"] = model_id

        if provider == "Custom":
            api_url = self.custom_api_url_var.get().strip()
            if not api_url:
                messagebox.showerror("错误", "自定义API地址不能为空！", parent=self)
                return None
            details["api_url"] = api_url

        return name, details

    def _add_model(self):
        data = self._collect_and_validate()
        if not data: return
        name, details = data
        if name in self.models:
            messagebox.showerror("错误", f"名为 '{name}' 的配置已存在！", parent=self)
            return
        self.models[name] = details
        self._save_and_update()

    def _update_model(self):
        selected = self.tree.selection()
        if not selected:
            messagebox.showerror("错误", "请选择要更新的模型！", parent=self)
            return
        original_name = selected[0]
        data = self._collect_and_validate()
        if not data: return
        name, details = data
        if original_name != name and name in self.models:
            messagebox.showerror("错误", f"新的配置名称 '{name}' 已存在！", parent=self)
            return
        if original_name != name:
            del self.models[original_name]
        self.models[name] = details
        self._save_and_update()
        if self.app_instance.current_model_name_var.get() == original_name:
            self.app_instance.current_model_name_var.set(name)
            self.app_instance.on_model_selected()

    def _delete_model(self):
        selected = self.tree.selection()
        if not selected: return
        name = selected[0]
        if messagebox.askyesno("确认删除", f"确定要删除配置 '{name}' 吗？", parent=self):
            del self.models[name]
            self._save_and_update()
            if self.app_instance.current_model_name_var.get() == name:
                self.app_instance.current_model_name_var.set("")
                self.app_instance.on_model_selected()

    def _save_and_update(self):
        self._load_models_to_treeview()
        self.app_instance.save_config(self.app_instance.config_file)
        self.app_instance.update_model_combobox()

    def _on_closing(self):
        self.grab_release()
        self.destroy()

# --- 主应用 TranslatorApp (已重构) ---
class TranslatorApp(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("通用AI翻译工具")
        self.geometry("1600x1000")
        self.style = ttk.Style(self)
        self.style.theme_use('clam')
        self.translator = None
        self.config_file = "config.json"
        
        self.models = {} 
        self.proxies = {}
        
        self.current_model_name_var = tk.StringVar()
        self.current_proxy_name_var = tk.StringVar(value="无代理")
        self.file_path_var = tk.StringVar()
        self.selected_src_display_var = tk.StringVar(value="未选择")
        self.selected_tgt_display_var = tk.StringVar(value="未选择")
        self.src_col_var = tk.StringVar()
        self.tgt_col_var = tk.StringVar()
        self.src_row_var = tk.StringVar()
        
        self._create_widgets()
        self.load_config(self.config_file) 
        
        text_handler = TextWidgetHandler(self.status_text)
        text_handler.setLevel(logging.DEBUG)
        text_formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        text_handler.setFormatter(text_formatter)
        logger.addHandler(text_handler)
        logger.info("应用程序启动成功。")

    def _create_widgets(self):
        main_pane = ttk.PanedWindow(self, orient=tk.HORIZONTAL)
        main_pane.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)
        control_panel_frame = ttk.Frame(main_pane)
        main_pane.add(control_panel_frame, weight=3)
        log_panel_frame = ttk.LabelFrame(main_pane, text="状态和日志", padding="10")
        main_pane.add(log_panel_frame, weight=2)
        control_panel_frame.rowconfigure(2, weight=1)
        control_panel_frame.columnconfigure(0, weight=1)
        
        file_preview_frame = ttk.LabelFrame(control_panel_frame, text="1. Excel文件预览与列选择", padding="10")
        file_preview_frame.grid(row=0, column=0, sticky="nsew", pady=5)
        file_preview_frame.columnconfigure(1, weight=1)
        file_preview_frame.rowconfigure(1, weight=1)
        ttk.Label(file_preview_frame, text="Excel文件:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        self.file_path_entry = ttk.Entry(file_preview_frame, textvariable=self.file_path_var, state='readonly')
        self.file_path_entry.grid(row=0, column=1, sticky=tk.EW, padx=5)
        ttk.Button(file_preview_frame, text="浏览...", command=self.browse_file).grid(row=0, column=2, padx=5)
        ttk.Label(file_preview_frame, text="源语言列/行: ").grid(row=2, column=0, sticky=tk.W, padx=5, pady=2)
        ttk.Label(file_preview_frame, textvariable=self.selected_src_display_var, foreground="green").grid(row=2, column=1, sticky=tk.W, padx=5, pady=2)
        ttk.Label(file_preview_frame, text="目标语言列: ").grid(row=3, column=0, sticky=tk.W, padx=5, pady=2)
        ttk.Label(file_preview_frame, textvariable=self.selected_tgt_display_var, foreground="blue").grid(row=3, column=1, sticky=tk.W, padx=5, pady=2)
        self.excel_preview = ExcelPreview(file_preview_frame, self)
        self.excel_preview.grid(row=1, column=0, columnspan=3, sticky=tk.NSEW, pady=5)

        lang_frame = ttk.LabelFrame(control_panel_frame, text="2. 语言设置", padding="10")
        lang_frame.grid(row=1, column=0, sticky="ew", pady=5)
        self.src_lang_var = tk.StringVar(value="俄语")
        self.tgt_lang_var = tk.StringVar(value="简体中文")
        ttk.Label(lang_frame, text="源语言:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        ttk.Entry(lang_frame, textvariable=self.src_lang_var).grid(row=0, column=1, padx=5)
        ttk.Label(lang_frame, text="目标语言:").grid(row=0, column=2, sticky=tk.W, padx=5, pady=5)
        ttk.Entry(lang_frame, textvariable=self.tgt_lang_var).grid(row=0, column=3, padx=5)
        
        api_proxy_frame = ttk.LabelFrame(control_panel_frame, text="3. AI与网络设置", padding="10")
        api_proxy_frame.grid(row=2, column=0, sticky="nsew", pady=5)
        api_proxy_frame.columnconfigure(1, weight=1)
        api_proxy_frame.rowconfigure(2, weight=1)

        ttk.Label(api_proxy_frame, text="网络代理:").grid(row=0, column=0, sticky=tk.W, padx=5, pady=5)
        self.proxy_combobox = ttk.Combobox(api_proxy_frame, textvariable=self.current_proxy_name_var, state="readonly")
        self.proxy_combobox.grid(row=0, column=1, sticky=tk.EW, padx=5)
        ttk.Button(api_proxy_frame, text="代理管理...", command=self.open_proxy_manager).grid(row=0, column=2, padx=5)
        
        ttk.Label(api_proxy_frame, text="AI模型配置:").grid(row=1, column=0, sticky=tk.W, padx=5, pady=5)
        self.model_combobox = ttk.Combobox(api_proxy_frame, textvariable=self.current_model_name_var, state="readonly")
        self.model_combobox.grid(row=1, column=1, sticky=tk.EW, padx=5)
        self.model_combobox.bind("<<ComboboxSelected>>", self.on_model_selected)
        ttk.Button(api_proxy_frame, text="模型管理...", command=self.open_model_manager).grid(row=1, column=2, padx=5)
        
        ttk.Label(api_proxy_frame, text="提示词模板:").grid(row=2, column=0, sticky=tk.NW, padx=5, pady=5)
        self.prompt_text = tk.Text(api_proxy_frame, height=12, wrap=tk.WORD)
        self.prompt_text.grid(row=2, column=1, columnspan=2, sticky=tk.NSEW, padx=5, pady=5)

        control_buttons_frame = ttk.Frame(control_panel_frame)
        control_buttons_frame.grid(row=3, column=0, sticky=tk.E, pady=10)
        self.start_button = ttk.Button(control_buttons_frame, text="开始翻译", command=self.start_translation)
        self.start_button.pack(side=tk.RIGHT, padx=5)
        ttk.Button(control_buttons_frame, text="保存配置", command=lambda: self.save_config(self.config_file)).pack(side=tk.LEFT, padx=5)
        
        log_panel_frame.rowconfigure(0, weight=1)
        log_panel_frame.columnconfigure(0, weight=1)
        self.status_text = tk.Text(log_panel_frame, wrap=tk.WORD, state=tk.DISABLED)
        self.status_text.grid(row=0, column=0, sticky=tk.NSEW)
        status_scroll = ttk.Scrollbar(log_panel_frame, orient="vertical", command=self.status_text.yview)
        status_scroll.grid(row=0, column=1, sticky="ns")
        self.status_text.configure(yscrollcommand=status_scroll.set)

    def start_translation(self):
        if not self.file_path_var.get() or not self.src_col_var.get() or not self.tgt_col_var.get() or not self.src_row_var.get():
            messagebox.showerror("错误", "请先选择Excel文件、源语言的起始单元格和目标语言列。")
            return
        model_name = self.current_model_name_var.get()
        if not model_name or model_name not in self.models:
            messagebox.showerror("错误", "请选择一个有效的AI模型配置。")
            return
        
        file_path = self.file_path_var.get()
        if not messagebox.askokcancel("开始翻译前确认", f"请确保您已关闭以下文件，否则可能导致保存失败:\n\n{os.path.basename(file_path)}\n\n点击“确定”开始翻译。"):
            return

        self.start_button.config(state=tk.DISABLED)
        logger.info("翻译任务启动...")
        
        threading.Thread(target=self._translation_worker, daemon=True).start()

    def _translation_worker(self):
        workbook = None
        file_path = self.file_path_var.get()
        try:
            model_details = self.models[self.current_model_name_var.get()]
            proxy_name = self.current_proxy_name_var.get()
            proxy_config = self.proxies.get(proxy_name) if proxy_name != "无代理" else None

            self.translator = translator.Translator(
                api_key=model_details.get("api_key"),
                model_id=model_details.get("model_id"),
                api_provider=model_details.get("provider"),
                custom_api_url=model_details.get("api_url"),
                proxy_config=proxy_config
            )
            
            workbook = openpyxl.load_workbook(file_path)
            sheet = workbook.active
            src_col_idx = openpyxl.utils.column_index_from_string(self.src_col_var.get())
            tgt_col_idx = openpyxl.utils.column_index_from_string(self.tgt_col_var.get())
            start_row = int(self.src_row_var.get())
            
            all_sources = []
            row_map = []
            for r_idx in range(start_row, sheet.max_row + 1):
                cell_value = sheet.cell(row=r_idx, column=src_col_idx).value
                all_sources.append(str(cell_value) if cell_value is not None else "")
                row_map.append(r_idx)

            if not any(s.strip() for s in all_sources):
                logger.info("在指定列中未找到需要翻译的文本。")
                self.after(0, lambda: messagebox.showinfo("完成", "未找到需要翻译的文本。"))
                return

            logger.info(f"共找到 {len(all_sources)} 行文本准备翻译。")

            prompt_template = self.prompt_text.get("1.0", tk.END)
            source_language = self.src_lang_var.get()
            target_language = self.tgt_lang_var.get()
            batch_size = 100
            all_results_valid = True

            for i in range(0, len(all_sources), batch_size):
                batch_sources = all_sources[i:i+batch_size]
                batch_row_map = row_map[i:i+batch_size]
                
                logger.info(f"正在处理批次 (行 {batch_row_map[0]}-{batch_row_map[-1]})...")

                translated_texts = self.translator.translate_batch(
                    batch_sources, 
                    prompt_template, 
                    source_language, 
                    target_language
                )

                if len(translated_texts) != len(batch_sources):
                    logger.error(f"批次翻译失败: 返回结果数量 ({len(translated_texts)}) 与源文本数量 ({len(batch_sources)}) 不匹配。")
                    all_results_valid = False
                    for original_row in batch_row_map:
                        sheet.cell(row=original_row, column=tgt_col_idx).value = "[批次翻译失败:行数不匹配]"
                else:
                    for j, original_row in enumerate(batch_row_map):
                        sheet.cell(row=original_row, column=tgt_col_idx).value = translated_texts[j]
                
                logger.info(f"批次 (行 {batch_row_map[0]}-{batch_row_map[-1]}) 已在内存中处理完成。")
                time.sleep(1)
            
            if all_results_valid:
                logger.info("所有批次处理完毕，准备保存文件...")
            else:
                logger.warning("部分批次翻译失败，请检查Excel文件中的错误信息。准备保存文件...")

            workbook.save(file_path)
            logger.info(f"所有翻译任务完成并成功保存到文件: {file_path}")
            self.after(0, lambda: messagebox.showinfo("完成", "所有翻译任务已完成！"))

        except Exception as e:
            logger.exception(f"翻译线程发生严重错误: {e}")
            self.after(0, lambda err=e: messagebox.showerror("错误", f"翻译过程中发生错误:\n{err}"))
        finally:
            if workbook: workbook.close()
            self.after(0, lambda: self.start_button.config(state=tk.NORMAL))

    def get_default_prompt(self):
        return f"You are an expert translator. Your task is to translate a batch of texts from {{source_language}} to {{target_language}}. The texts are separated by a unique delimiter: '{translator.LINE_SEPARATOR}'.\n\n**CRITICAL INSTRUCTIONS:**\n1.  Translate each segment of text between the delimiters individually.\n2.  You MUST preserve the exact same delimiter '{translator.LINE_SEPARATOR}' between each translated segment.\n3.  The number of delimiters in your output MUST be exactly one less than the number of text segments in the input.\n4.  If a segment in the input is empty or contains only whitespace, you MUST output an empty segment in its place, followed by the delimiter.\n5.  Do NOT add any extra text, explanations, or formatting. Your response should only contain the translated texts separated by the specified delimiter.\n\n**EXAMPLE:**\n- **INPUT TEXT:**\nHello world{translator.LINE_SEPARATOR}{translator.LINE_SEPARATOR}How are you?\n- **EXPECTED OUTPUT (to Spanish):**\nHola mundo{translator.LINE_SEPARATOR}{translator.LINE_SEPARATOR}¿Cómo estás?\n\n--- TEXT TO TRANSLATE ---\n{{text_to_translate}}"

    def save_config(self, file_path):
        config_data = {
            "models": self.models,
            "proxies": self.proxies,
            "current_model_name": self.current_model_name_var.get(),
            "current_proxy_name": self.current_proxy_name_var.get(),
            "source_language": self.src_lang_var.get(),
            "target_language": self.tgt_lang_var.get(),
            "prompt_template": self.prompt_text.get("1.0", tk.END).strip(),
            "src_col": self.src_col_var.get(),
            "tgt_col": self.tgt_col_var.get(),
            "src_row": self.src_row_var.get()
        }
        try:
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(config_data, f, ensure_ascii=False, indent=4)
            logger.info(f"配置已保存到 {file_path}")
        except Exception as e:
            logger.error(f"保存配置失败: {e}")

    def load_config(self, file_path):
        default_prompt = self.get_default_prompt()
        if not os.path.exists(file_path):
            logger.info("未找到配置文件，创建默认配置。")
            self.models["Gemini (请先配置)"] = {
                "provider": "Gemini",
                "api_key": "在此处粘贴您的Google AI API密钥",
                "model_id": ""
            }
            self.models["DeepSeek (请先配置)"] = {
                "provider": "DeepSeek",
                "api_key": "在此处粘贴您的DeepSeek API密钥",
                "model_id": "deepseek-chat"
            }
            self.proxies = {}
            self.prompt_text.insert("1.0", default_prompt)
            self.save_config(file_path)
            config_data = {}
        else:
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    config_data = json.load(f)
                self.models = config_data.get("models", {})
                self.proxies = config_data.get("proxies", {})
            except Exception as e:
                logger.error(f"加载配置失败: {e}")
                config_data = {}
                self.models = {}
                self.proxies = {}

        self.update_model_combobox()
        self.update_proxy_combobox()

        self.current_model_name_var.set(config_data.get("current_model_name", ""))
        self.current_proxy_name_var.set(config_data.get("current_proxy_name", "无代理"))
        
        self.src_lang_var.set(config_data.get("source_language", "俄语"))
        self.tgt_lang_var.set(config_data.get("target_language", "简体中文"))
        
        prompt_from_config = config_data.get("prompt_template", "").strip()
        self.prompt_text.delete("1.0", tk.END)
        self.prompt_text.insert("1.0", prompt_from_config if prompt_from_config else default_prompt)
        
        self.src_col_var.set(config_data.get("src_col", ""))
        self.tgt_col_var.set(config_data.get("tgt_col", ""))
        self.src_row_var.set(config_data.get("src_row", ""))
        
        self.on_model_selected()
        self.update_selection_display()

    def open_model_manager(self): ModelManagerWindow(self, self)
    def open_proxy_manager(self): ProxyManagerWindow(self, self)
    
    def on_model_selected(self, event=None):
        model_name = self.current_model_name_var.get()
        logger.info(f"AI模型配置已选择: '{model_name}'")

    def update_model_combobox(self):
        model_names = list(self.models.keys())
        self.model_combobox['values'] = model_names
        if self.current_model_name_var.get() not in model_names:
            self.current_model_name_var.set(model_names[0] if model_names else "")
    
    def update_proxy_combobox(self):
        proxy_names = ["无代理"] + list(self.proxies.keys())
        self.proxy_combobox['values'] = proxy_names
        if self.current_proxy_name_var.get() not in proxy_names:
            self.current_proxy_name_var.set("无代理")

    def browse_file(self):
        file_path = filedialog.askopenfilename(filetypes=(("Excel files", "*.xlsx *.xls"), ("All files", "*.*的发展")))
        if file_path:
            self.file_path_var.set(file_path)
            try:
                workbook = openpyxl.load_workbook(file_path, data_only=True)
                self.excel_preview.load_sheet(workbook.active)
            except Exception as e:
                messagebox.showerror("错误", f"无法加载Excel文件:\n{e}")

    def update_selection_display(self):
        src_col, src_row = self.excel_preview.get_selected_source_coords()
        tgt_col = self.excel_preview.get_selected_target_col()
        self.selected_src_display_var.set(f"列: {src_col}, 行: {src_row}" if src_col and src_row else "未选择")
        self.selected_tgt_display_var.set(f"列: {tgt_col}" if tgt_col else "未选择")
        if src_col and src_row:
            self.src_col_var.set(src_col)
            self.src_row_var.set(str(src_row))
        if tgt_col: self.tgt_col_var.set(tgt_col)

if __name__ == "__main__":
    app = TranslatorApp()
    app.mainloop()
